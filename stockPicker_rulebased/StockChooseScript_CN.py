import random
import time
from collections import Counter

import numpy as np
import pandas as pd
import requests
import yfinance as yf

print("CN Stock Screener - Fund Flow + Technical Filters")

# ------------------------------
# Configuration
# ------------------------------
LOOKBACK_DAYS = 120 # 策略阈值
MIN_RETURN_5D = 0.03 #
MIN_RETURN_7D_CUM = 0.25
MIN_VOLUME_INCREASE = 1.2
MIN_ROE_PCT = 10.0
ALLOW_MISSING_ROE = True
MAX_PRICE_RATIO = 0.90
CLOSE_ABOVE_MA5_DAYS = 15
CLOSE_ABOVE_MA5_MIN = 11
MA_SLOPE_DAYS = 5
MIN_APPEARANCES_7D = 2
LAST_TRADE_DAYS = 10
DEBUG_FILTER = True

EASTMONEY_URL = "https://push2.eastmoney.com/api/qt/clist/get"
EASTMONEY_UT = "8dec03ba335b81bf4ebdf7b29ec27d15"
FIELDS_STAT_1 = (
    "f12,f14,f2,f3,f62,f184,f66,f69,f72,f75,f78,f81,f84,f87,f204,f205,f124,f1,f13"
) # 括号内多element表示元组， 现在还是字符
FS_HSA = "m:0+t:6+f:!2,m:0+t:13+f:!2,m:0+t:80+f:!2,m:1+t:2+f:!2,m:1+t:23+f:!2"

CACHE_PATH = "fund_flow_top10_cache.csv"
FUND_FLOW_HISTORY_CSV_PATH = "fund_flow_top10_history.csv"
FUND_FLOW_HISTORY_XLSX_PATH = "fund_flow_top10_history.xlsx"
SELECTED_CSV_HISTORY_PATH = "selected_cn_history.csv"
SELECTED_XLSX_HISTORY_PATH = "selected_cn_history.xlsx"

COLUMN_MAP = {
    "\u65e5\u671f": "date",
    "\u5f00\u76d8": "open",
    "\u6536\u76d8": "close",
    "\u6700\u9ad8": "high",
    "\u6700\u4f4e": "low",
    "\u6210\u4ea4\u91cf": "volume",
}

# 获取最近 n 个交易日日期列表
YF_PRICE_COLUMN_MAP = {
    "Date": "date",
    "Datetime": "date",
    "Open": "open",
    "Close": "close",
    "High": "high",
    "Low": "low",
    "Volume": "volume",
}


def to_yf_symbol(code):
    code = str(code).zfill(6)
    if code.startswith("6"):
        suffix = "SS"
    elif code.startswith(("0", "3")):
        suffix = "SZ"
    elif code.startswith(("4", "8", "9")):
        suffix = "BJ"
    else:
        suffix = "SZ"
    return f"{code}.{suffix}"


def get_last_trade_dates(n):
    try:
        cal = yf.Ticker("000001.SS").history(
            period=f"{max(60, n * 6)}d",
            interval="1d",
            auto_adjust=False,
            actions=False,
        )
        if cal is None or cal.empty:
            dates = []
        else:
            idx = pd.DatetimeIndex(cal.index)
            if idx.tz is not None:
                idx = idx.tz_convert(None)
            dates = idx.strftime("%Y-%m-%d").tolist()
        if dates:  # when date is empty return false
            return dates[-n:]
    except Exception:
        pass
    return pd.bdate_range(end=pd.Timestamp.today(), periods=n).strftime("%Y-%m-%d").tolist()


def fetch_price_history_yf(code, start_date="2018-01-01"):
    symbol = to_yf_symbol(code)
    df = yf.Ticker(symbol).history(
        start=start_date,
        interval="1d",
        auto_adjust=False,
        actions=False,
    )
    if df is None or df.empty:
        return pd.DataFrame()
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = [c[0] if isinstance(c, tuple) else c for c in df.columns]
    return df.reset_index().rename(columns=YF_PRICE_COLUMN_MAP)


def fetch_fund_flow_top10():
    params = {
        "pn": 1,
        "pz": 10,
        "po": 1,
        "np": 1,
        "fltt": 2,
        "invt": 2,
        "fid": "f62",
        "fs": FS_HSA,
        "fields": FIELDS_STAT_1,
        "ut": EASTMONEY_UT,
    }
    resp = requests.get(EASTMONEY_URL, params=params, timeout=15)
    resp.raise_for_status()
    payload = resp.json()
    items = payload.get("data", {}).get("diff", []) or []
    results = []
    for idx, item in enumerate(items, start=1):
        code = item.get("f12")
        name = item.get("f14")
        if code:
            code = str(code).zfill(6)
            results.append({"code": code, "name": name, "rank": idx})
    return results


def load_cache():
    try:
        df = pd.read_csv(CACHE_PATH, dtype={"code": str})
        df["code"] = df["code"].astype(str).str.zfill(6)
        if "trade_date" in df.columns:
            df["trade_date"] = (
                pd.to_datetime(df["trade_date"], errors="coerce")
                .dt.strftime("%Y-%m-%d")
            )
        return df
    except Exception:
        return pd.DataFrame(columns=["trade_date", "code", "name", "rank"])


def save_cache(df):
    try:
        df.to_csv(CACHE_PATH, index=False, encoding="utf-8-sig")
    except PermissionError:
        alt_path = f"fund_flow_top10_cache_{time.strftime('%Y%m%d_%H%M%S')}.csv"
        df.to_csv(alt_path, index=False, encoding="utf-8-sig")
        print(
            f"Warning: no permission to write {CACHE_PATH}. "
            f"Wrote cache to {alt_path} instead. Close any app using the file."
        )


def update_cache_for_date(trade_date):
    cache_df = load_cache()
    existing = cache_df[cache_df["trade_date"] == trade_date]
    if len(existing) >= 10:
        append_history_csv(existing, FUND_FLOW_HISTORY_CSV_PATH, code_columns=["code"])
        append_history_excel(
            existing,
            FUND_FLOW_HISTORY_XLSX_PATH,
            sheet_name="History",
            code_columns=["code"],
        )
        return cache_df
    rows = fetch_fund_flow_top10()
    if not rows:
        print(f"Warning: fund flow list is empty for {trade_date}.")
        return cache_df
    new_df = pd.DataFrame(rows)
    new_df["trade_date"] = trade_date
    merged = pd.concat([cache_df, new_df], ignore_index=True)
    save_cache(merged)
    append_history_csv(new_df, FUND_FLOW_HISTORY_CSV_PATH, code_columns=["code"])
    append_history_excel(
        new_df,
        FUND_FLOW_HISTORY_XLSX_PATH,
        sheet_name="History",
        code_columns=["code"],
    )
    return merged


def get_candidates_from_cache(trade_dates):
    cache_df = load_cache()
    available_dates = sorted(set(cache_df["trade_date"]).intersection(trade_dates))
    missing = [d for d in trade_dates if d not in set(cache_df["trade_date"])]
    if missing:
        print(
            "Warning: missing cached fund flow dates: "
            + ", ".join(missing)
            + ". Run this script on each trading day to build history."
        )
    recent = cache_df[cache_df["trade_date"].isin(available_dates)]
    counts = Counter(recent["code"].tolist())
    available_days = len(available_dates)
    if available_days == 0:
        return {}
    if available_days < 3:
        threshold = 1
        print(
            f"Only {available_days} cached day(s). Using >=1 rule: "
            f"appearances >= {threshold}."
        )
    else:
        threshold = MIN_APPEARANCES_7D
    return {code: cnt for code, cnt in counts.items() if cnt >= threshold}


def write_excel_with_text_code(df, path, sheet_name="data"):
    try:
        import openpyxl  # noqa: F401
        from openpyxl.styles import numbers
    except Exception:
        print(f"Warning: openpyxl not available, skipped Excel output: {path}")
        return
    df_out = df.copy()
    for col in ["code", "Code"]:
        if col in df_out.columns:
            df_out[col] = df_out[col].astype(str).str.zfill(6)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df_out.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.book[sheet_name]
        header = [cell.value for cell in ws[1]]
        for col_name in ["code", "Code"]:
            if col_name in header:
                col_idx = header.index(col_name) + 1
                for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        cell.number_format = numbers.FORMAT_TEXT


def load_history_excel(path, sheet_name):
    try:
        df = pd.read_excel(path, sheet_name=sheet_name, dtype={"code": str, "Code": str})
        for col in ["code", "Code"]:
            if col in df.columns:
                df[col] = df[col].astype(str).str.zfill(6)
        return df
    except Exception:
        return pd.DataFrame()


def append_history_excel(df, path, sheet_name, code_columns):
    try:
        import openpyxl  # noqa: F401
    except Exception:
        print(f"Warning: openpyxl not available, skipped Excel output: {path}")
        return
    df_out = df.copy()
    df_out["run_time"] = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")
    for col in code_columns:
        if col in df_out.columns:
            df_out[col] = df_out[col].astype(str).str.zfill(6)
    history = load_history_excel(path, sheet_name)
    history = pd.concat([history, df_out], ignore_index=True)
    write_excel_with_text_code(history, path, sheet_name=sheet_name)


def append_history_csv(df, path, code_columns):
    df_out = df.copy()
    df_out["run_time"] = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")
    for col in code_columns:
        if col in df_out.columns:
            df_out[col] = df_out[col].astype(str).str.zfill(6)
    try:
        history = pd.read_csv(path, dtype={col: str for col in code_columns})
        for col in code_columns:
            if col in history.columns:
                history[col] = history[col].astype(str).str.zfill(6)
    except Exception:
        history = pd.DataFrame()
    history = pd.concat([history, df_out], ignore_index=True)
    history.to_csv(path, index=False, encoding="utf-8-sig")


def write_levels_excel(df, path, mode="A"):
    try:
        import openpyxl  # noqa: F401
    except Exception:
        print(f"Warning: openpyxl not available, skipped Excel output: {path}")
        return
    import os
    run_time = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")
    df_run = df.copy()
    df_run["run_time"] = run_time
    for col in ["Code"]:
        if col in df_run.columns:
            df_run[col] = df_run[col].astype(str).str.zfill(6)

    mode_key = str(mode).strip().upper() or "A"
    if mode_key == "A":
        mode_suffix = "7cacheData"
    elif mode_key == "B":
        mode_suffix = "onetime"
    else:
        mode_suffix = mode_key
    history_sheet = f"History_{mode_suffix}"
    summary_sheet = f"Summary_{mode_suffix}"
    level_sheet_fmt = f"Level_{{level}}_{mode_suffix}"

    history = load_history_excel(path, history_sheet)
    history = pd.concat([history, df_run], ignore_index=True)

    summary_rows = []
    for level in [1, 2, 3]:
        codes = df_run[df_run["Level"] == level]["Code"].tolist()
        summary_rows.append({"Level": level, "Codes": ", ".join(codes) if codes else ""})
    summary_df = pd.DataFrame(summary_rows)
    summary_history = load_history_excel(path, summary_sheet)
    summary_history = pd.concat([summary_history, summary_df], ignore_index=True)

    writer_kwargs = {"engine": "openpyxl"}
    if os.path.exists(path):
        writer_kwargs["mode"] = "a"
        writer_kwargs["if_sheet_exists"] = "replace"

    with pd.ExcelWriter(path, **writer_kwargs) as writer:
        history.to_excel(writer, index=False, sheet_name=history_sheet)
        for level in [1, 2, 3]:
            sheet_name = level_sheet_fmt.format(level=level)
            df_level = df_run[df_run["Level"] == level].copy()
            if df_level.empty:
                df_level = pd.DataFrame(columns=df_run.columns)
            df_level = df_level.sort_values(["Appearances_7D"], ascending=[False])
            level_history = load_history_excel(path, sheet_name)
            level_history = pd.concat([level_history, df_level], ignore_index=True)
            level_history.to_excel(writer, index=False, sheet_name=sheet_name)
        summary_history.to_excel(writer, index=False, sheet_name=summary_sheet)

    wb = openpyxl.load_workbook(path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header = [cell.value for cell in ws[1]]
        for col_name in ["code", "Code"]:
            if col_name in header:
                col_idx = header.index(col_name) + 1
                for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        cell.number_format = openpyxl.styles.numbers.FORMAT_TEXT
    wb.save(path)


def parse_roe_value(value):
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return None
    if isinstance(value, str):
        value = value.replace("%", "").replace(",", "").strip()
    try:
        val = float(value)
    except Exception:
        return None
    if val <= 1.0:
        val *= 100.0
    return val


def get_latest_roe_pct(code):
    symbol = to_yf_symbol(code)
    ticker = yf.Ticker(symbol)
    try:
        info = ticker.get_info() if hasattr(ticker, "get_info") else ticker.info
    except Exception:
        return None
    if not isinstance(info, dict):
        return None
    for key in ("returnOnEquity", "returnOnEquityTTM"):
        if info.get(key) is not None:
            return parse_roe_value(info.get(key))
    return None


def ma_divergence_up(df):
    if len(df) < MA_SLOPE_DAYS + 1:
        return False
    m7 = df["MA7"]
    m14 = df["MA14"]
    m21 = df["MA21"]
    if not (m7.iloc[-1] > m14.iloc[-1] > m21.iloc[-1]):
        return False
    if not (
        m7.iloc[-1] > m7.iloc[-1 - MA_SLOPE_DAYS]
        and m14.iloc[-1] > m14.iloc[-1 - MA_SLOPE_DAYS]
        and m21.iloc[-1] > m21.iloc[-1 - MA_SLOPE_DAYS]
    ):
        return False
    gap_now_7_14 = m7.iloc[-1] - m14.iloc[-1]
    gap_now_14_21 = m14.iloc[-1] - m21.iloc[-1]
    gap_prev_7_14 = m7.iloc[-1 - MA_SLOPE_DAYS] - m14.iloc[-1 - MA_SLOPE_DAYS]
    gap_prev_14_21 = m14.iloc[-1 - MA_SLOPE_DAYS] - m21.iloc[-1 - MA_SLOPE_DAYS]
    return gap_now_7_14 > gap_prev_7_14 and gap_now_14_21 > gap_prev_14_21


def close_above_ma5(df):
    if len(df) < CLOSE_ABOVE_MA5_DAYS:
        return False
    recent_close = df["close"].tail(CLOSE_ABOVE_MA5_DAYS)
    recent_ma5 = df["MA5"].tail(CLOSE_ABOVE_MA5_DAYS)
    return (recent_close > recent_ma5).sum() >= CLOSE_ABOVE_MA5_MIN


def screen_stock(code, appearances_7d):
    df = fetch_price_history_yf(code, start_date="2018-01-01")
    if df is None or df.empty:
        return None, {"skip": "empty_hist"}

    df = df.rename(columns=COLUMN_MAP)
    if "date" not in df.columns:
        return None, {"skip": "missing_date_col"}

    df["date"] = pd.to_datetime(df["date"])
    df = df.sort_values("date").reset_index(drop=True)

    if len(df) < LOOKBACK_DAYS:
        return None, {"skip": f"insufficient_days<{LOOKBACK_DAYS}"}

    for col in ["close", "volume"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    df["MA5"] = df["close"].rolling(5).mean()
    df["MA7"] = df["close"].rolling(7).mean()
    df["MA10"] = df["close"].rolling(10).mean()
    df["MA14"] = df["close"].rolling(14).mean()
    df["MA20"] = df["close"].rolling(20).mean()
    df["MA21"] = df["close"].rolling(21).mean()
    df["return_5d"] = df["close"].pct_change(5)
    df["return_7d"] = df["close"].pct_change(7)

    volume_now = df["volume"].tail(5).mean()
    volume_prev = df["volume"].tail(15).head(10).mean()
    volume_ratio = volume_now / volume_prev if volume_prev else 0

    highest_price = df["close"].max()
    current_price = df["close"].iloc[-1]
    price_ratio = current_price / highest_price if highest_price else 0

    roe_pct = get_latest_roe_pct(code)

    p1_ma = ma_divergence_up(df)
    p1_close = close_above_ma5(df)
    p1 = p1_ma and p1_close
    p2 = df["return_7d"].iloc[-1] > MIN_RETURN_7D_CUM
    roe_check_pass = (
        True if (roe_pct is None and ALLOW_MISSING_ROE) else (roe_pct is not None and roe_pct >= MIN_ROE_PCT)
    )
    p3_checks = {
        "roe": roe_check_pass,
        "ma_trend": df["MA5"].iloc[-1] > df["MA10"].iloc[-1] > df["MA20"].iloc[-1],
        "return_5d": df["return_5d"].iloc[-1] > MIN_RETURN_5D,
        "volume": volume_ratio >= MIN_VOLUME_INCREASE,
        "price_ratio": price_ratio < MAX_PRICE_RATIO,
    }
    p3 = all(p3_checks.values())

    level = None
    if p1:
        level = 1
    if p1 and p2:
        level = 2
    if p1 and p2 and p3:
        level = 3
    debug = {
        "p1": p1,
        "p1_ma": p1_ma,
        "p1_close": p1_close,
        "p2": p2,
        "p3": p3,
        "p3_failed": [k for k, v in p3_checks.items() if not v],
        "roe_missing_skipped": roe_pct is None and ALLOW_MISSING_ROE,
        "roe_pct": roe_pct,
        "return_5d": df["return_5d"].iloc[-1],
        "return_7d": df["return_7d"].iloc[-1],
        "volume_ratio": volume_ratio,
        "price_ratio": price_ratio,
    }
    if level is None:
        return None, debug

    roe_out = round(roe_pct, 2) if roe_pct is not None else None
    return {
        "Code": code,
        "ROE_Pct": roe_out,
        "5D_Return": round(df["return_5d"].iloc[-1], 3),
        "7D_Return": round(df["return_7d"].iloc[-1], 3),
        "Volume_Ratio": round(volume_ratio, 2),
        "Price/High": round(price_ratio, 2),
        "Appearances_7D": appearances_7d,
        "Level": level,
    }, debug


mode = input("Select mode: A=cache history, B=today only: ").strip().upper()
if mode not in {"A", "B"}:
    print("Unknown mode, defaulting to A (cache history).")
    mode = "A"

trade_dates = get_last_trade_dates(LAST_TRADE_DAYS)
latest_trade_date = trade_dates[-1]
cache_df = update_cache_for_date(latest_trade_date)
if mode == "B":
    today_rows = fetch_fund_flow_top10()
    candidate_counts = {row["code"]: 1 for row in today_rows} if today_rows else {}
else:
    candidate_counts = get_candidates_from_cache(trade_dates)

print(f"Trade dates (last {LAST_TRADE_DAYS}): {', '.join(trade_dates)}")
print(f"Candidates after fund flow filter: {len(candidate_counts)}")

results = []
for idx, (code, cnt) in enumerate(candidate_counts.items(), start=1):
    try:
        print(f"[{idx}/{len(candidate_counts)}] Checking {code} ...")
        hit, debug = screen_stock(code, cnt)
        if hit:
            print(f"Hit: {code}")
            results.append(hit)
        elif DEBUG_FILTER:
            if debug.get("skip"):
                print(f"  Skip {code}: {debug['skip']}")
            else:
                print(
                    "  Fail {code}: p1={p1} (ma={p1_ma}, close={p1_close}) "
                    "p2={p2} p3={p3} miss={miss} "
                    "r5={r5:.3f} r7={r7:.3f} vol={vol:.2f} pr={pr:.2f} roe={roe}".format(
                        code=code,
                        p1=debug["p1"],
                        p1_ma=debug["p1_ma"],
                        p1_close=debug["p1_close"],
                        p2=debug["p2"],
                        p3=debug["p3"],
                        miss=",".join(debug["p3_failed"]),
                        r5=debug["return_5d"],
                        r7=debug["return_7d"],
                        vol=debug["volume_ratio"],
                        pr=debug["price_ratio"],
                        roe=debug["roe_pct"],
                    )
                )
        time.sleep(random.uniform(0.3, 0.6))
    except Exception as exc:
        print(f"Error for {code}: {exc}")
        time.sleep(1)

df_res = pd.DataFrame(results)
print("\nSelected stocks by level:\n")
if df_res.empty:
    print("No stocks matched the rules.")
else:
    df_res = df_res.sort_values(["Level", "Appearances_7D"], ascending=[True, False])
    level1_codes = df_res[df_res["Level"] == 1]["Code"].tolist()
    level2_codes = df_res[df_res["Level"] == 2]["Code"].tolist()
    level3_codes = df_res[df_res["Level"] == 3]["Code"].tolist()
    print(f"满足优先级1: {', '.join(level1_codes) if level1_codes else 'None'}")
    print(
        f"满足优先级1&2: {', '.join(level2_codes) if level2_codes else 'None'}"
    )
    print(
        f"满足优先级3: {', '.join(level3_codes) if level3_codes else 'None'}"
    )

if not df_res.empty:
    append_history_csv(df_res, SELECTED_CSV_HISTORY_PATH, code_columns=["Code"])
    write_levels_excel(df_res, SELECTED_XLSX_HISTORY_PATH, mode=mode)
    print(f"Saved to {SELECTED_CSV_HISTORY_PATH}")

print("\nScan completed.")
